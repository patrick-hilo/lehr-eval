from io import BytesIO
from pathlib import Path

from fastapi.testclient import TestClient
from openpyxl import load_workbook

from lehr_eval.app import create_app
from lehr_eval.db import connect
from lehr_eval.exports import (
    ExportNotAvailable,
    ExportNotFound,
    build_single_export,
    build_teacher_export,
)
from lehr_eval.migrations import initialize_database


def test_single_export_contains_header_and_item_aggregates(tmp_path: Path):
    db_path, evaluation_id, _teacher_id = closed_evaluation_with_aggregates(tmp_path)

    content = build_single_export(db_path, evaluation_id)
    workbook = load_workbook(BytesIO(content))
    sheet = workbook.active

    assert sheet["A1"].value == "Schuljahr"
    assert sheet["B1"].value == "2025/26"
    assert sheet["A2"].value == "Klasse/Lerngruppe"
    assert sheet["B2"].value == "8b"
    assert sheet["A3"].value == "Klassenstufe"
    assert sheet["B3"].value == 8
    assert sheet["A4"].value == "Fach"
    assert sheet["B4"].value == "Mathematik"
    assert sheet["A5"].value == "Lehrkraft"
    assert sheet["B5"].value == "Frau Mueller"
    assert sheet["A6"].value == "Fragebogen-Version"
    assert sheet["B6"].value == "oberstufe-v1"
    assert sheet["A7"].value == "Erwartete Teilnehmerzahl"
    assert sheet["B7"].value == 24
    assert sheet["A8"].value == "tatsaechliche Beitritte"
    assert sheet["B8"].value == 3
    assert "Auswertung enthaelt nur aggregierte Daten" in sheet["A9"].value
    assert sheet["A10"].value == "Abschlussdatum"
    assert sheet["B10"].value == "2026-05-15 14:30:00"
    assert sheet["A12"].value == "Item"
    assert sheet["B12"].value == "0"
    assert sheet["C12"].value == "1"
    assert sheet["D12"].value == "2"
    assert sheet["E12"].value == "3"
    assert sheet["F12"].value == "Fehlend"
    assert sheet["G12"].value == "Beitritte"
    assert sheet["H12"].value == "Mittelwert"
    assert sheet["A13"].value == "Item 1"
    assert sheet["B13"].value == 1
    assert sheet["C13"].value == 0
    assert sheet["D13"].value == 1
    assert sheet["E13"].value == 1
    assert sheet["F13"].value == 0
    assert sheet["G13"].value == 3
    assert sheet["H13"].value == 2.0


def test_teacher_export_has_one_sheet_per_evaluation_and_no_overview(tmp_path: Path):
    db_path, first_id, teacher_id = closed_evaluation_with_aggregates(tmp_path)
    second_id = insert_evaluation(
        db_path,
        teacher_id=teacher_id,
        class_group="7a",
        subject="Deutsch",
        student_token="student-token-2",
        teacher_token="teacher-token-2",
    )
    insert_aggregate(db_path, second_id, "Item 1", joined_count=2)

    content = build_teacher_export(db_path, teacher_id, "2025/26")
    workbook = load_workbook(BytesIO(content))

    assert workbook.sheetnames == ["8b Mathematik", "7a Deutsch"]
    assert workbook["8b Mathematik"]["B1"].value == "2025/26"
    assert workbook["7a Deutsch"]["B4"].value == "Deutsch"
    assert "Overview" not in workbook.sheetnames
    assert len(workbook.sheetnames) == 2
    assert first_id != second_id


def test_teacher_export_numbers_sheet_name_collisions(tmp_path: Path):
    db_path, _first_id, teacher_id = closed_evaluation_with_aggregates(tmp_path)
    second_id = insert_evaluation(
        db_path,
        teacher_id=teacher_id,
        class_group="8/b",
        subject="Mathematik",
        student_token="student-token-2",
        teacher_token="teacher-token-2",
    )
    insert_aggregate(db_path, second_id, "Item 1", joined_count=1)

    content = build_teacher_export(db_path, teacher_id, "2025/26")
    workbook = load_workbook(BytesIO(content))

    assert workbook.sheetnames == ["8b Mathematik", "8b Mathematik 2"]


def test_teacher_export_numbers_case_insensitive_sheet_name_collisions(
    tmp_path: Path,
):
    db_path, _first_id, teacher_id = closed_evaluation_with_aggregates(tmp_path)
    second_id = insert_evaluation(
        db_path,
        teacher_id=teacher_id,
        class_group="8B",
        subject="Mathematik",
        student_token="student-token-2",
        teacher_token="teacher-token-2",
    )
    insert_aggregate(db_path, second_id, "Item 1", joined_count=1)

    content = build_teacher_export(db_path, teacher_id, "2025/26")
    workbook = load_workbook(BytesIO(content))

    assert workbook.sheetnames == ["8b Mathematik", "8B Mathematik 2"]
    assert all(len(name) <= 31 for name in workbook.sheetnames)


def test_export_does_not_include_teacher_email(tmp_path: Path):
    db_path, evaluation_id, _teacher_id = closed_evaluation_with_aggregates(tmp_path)

    content = build_single_export(db_path, evaluation_id)
    workbook = load_workbook(BytesIO(content), read_only=True)
    values = [
        value
        for sheet in workbook.worksheets
        for row in sheet.iter_rows(values_only=True)
        for value in row
        if value is not None
    ]

    assert "mueller@example.test" not in "\n".join(str(value) for value in values)


def test_export_escapes_formula_like_metadata(tmp_path: Path):
    db_path, evaluation_id, _teacher_id = closed_evaluation_with_aggregates(tmp_path)
    with connect(db_path) as db:
        db.execute("update teachers set name = '=HYPERLINK(\"http://example.test\")'")
        db.execute(
            """
            update evaluations
            set class_group = '+8b',
                subject = '-Mathematik',
                school_year = '@2025/26'
            where id = ?
            """,
            (evaluation_id,),
        )

    content = build_single_export(db_path, evaluation_id)
    workbook = load_workbook(BytesIO(content), data_only=False)
    sheet = workbook.active

    assert sheet["B1"].data_type == "s"
    assert sheet["B1"].value == "'@2025/26"
    assert sheet["B2"].value == "'+8b"
    assert sheet["B4"].value == "'-Mathematik"
    assert sheet["B5"].value == "'=HYPERLINK(\"http://example.test\")"


def test_single_export_rejects_non_closed_evaluation(tmp_path: Path):
    db_path, evaluation_id, _teacher_id = closed_evaluation_with_aggregates(tmp_path)
    with connect(db_path) as db:
        db.execute(
            "update evaluations set status = 'active' where id = ?", (evaluation_id,)
        )

    try:
        build_single_export(db_path, evaluation_id)
    except ExportNotAvailable:
        pass
    else:
        raise AssertionError("expected non-closed evaluation to be unavailable")


def test_teacher_export_includes_only_closed_evaluations(tmp_path: Path):
    db_path, _first_id, teacher_id = closed_evaluation_with_aggregates(tmp_path)
    active_id = insert_evaluation(
        db_path,
        teacher_id=teacher_id,
        class_group="7a",
        subject="Deutsch",
        student_token="student-token-2",
        teacher_token="teacher-token-2",
        status="active",
    )
    insert_aggregate(db_path, active_id, "Item 1", joined_count=2)

    content = build_teacher_export(db_path, teacher_id, "2025/26")
    workbook = load_workbook(BytesIO(content))

    assert workbook.sheetnames == ["8b Mathematik"]


def test_teacher_export_rejects_when_no_closed_evaluations_exist(tmp_path: Path):
    db_path, _first_id, teacher_id = closed_evaluation_with_aggregates(tmp_path)
    with connect(db_path) as db:
        db.execute("update evaluations set status = 'active'")

    try:
        build_teacher_export(db_path, teacher_id, "2025/26")
    except ExportNotFound:
        pass
    else:
        raise AssertionError("expected teacher export without closed evaluations to fail")


def test_admin_export_endpoint_requires_login_and_returns_xlsx(tmp_path: Path):
    db_path, evaluation_id, _teacher_id = closed_evaluation_with_aggregates(tmp_path)
    client = TestClient(create_app(db_path=db_path, admin_password="secret"))

    unauthenticated = client.get(
        f"/admin/evaluations/{evaluation_id}/export.xlsx", follow_redirects=False
    )
    assert unauthenticated.status_code == 303
    assert unauthenticated.headers["location"] == "/admin/login"

    login(client)
    response = client.get(f"/admin/evaluations/{evaluation_id}/export.xlsx")
    workbook = load_workbook(BytesIO(response.content))

    assert response.status_code == 200
    assert response.headers["content-type"] == (
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    assert workbook.active["B2"].value == "8b"


def test_admin_export_endpoint_rejects_non_closed_evaluation(tmp_path: Path):
    db_path, evaluation_id, _teacher_id = closed_evaluation_with_aggregates(tmp_path)
    with connect(db_path) as db:
        db.execute(
            "update evaluations set status = 'active' where id = ?", (evaluation_id,)
        )
    client = TestClient(create_app(db_path=db_path, admin_password="secret"))

    login(client)
    response = client.get(f"/admin/evaluations/{evaluation_id}/export.xlsx")

    assert response.status_code == 409


def test_admin_teacher_export_endpoint_returns_xlsx_after_login(tmp_path: Path):
    db_path, _evaluation_id, teacher_id = closed_evaluation_with_aggregates(tmp_path)
    client = TestClient(create_app(db_path=db_path, admin_password="secret"))

    login(client)
    response = client.get(f"/admin/teachers/{teacher_id}/2025%2F26/export.xlsx")
    workbook = load_workbook(BytesIO(response.content))

    assert response.status_code == 200
    assert response.headers["content-type"] == (
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    assert workbook.sheetnames == ["8b Mathematik"]


def test_admin_teacher_export_endpoint_returns_404_without_closed_evaluations(
    tmp_path: Path,
):
    db_path, _evaluation_id, teacher_id = closed_evaluation_with_aggregates(tmp_path)
    with connect(db_path) as db:
        db.execute("update evaluations set status = 'active'")
    client = TestClient(create_app(db_path=db_path, admin_password="secret"))

    login(client)
    response = client.get(f"/admin/teachers/{teacher_id}/2025%2F26/export.xlsx")

    assert response.status_code == 404


def test_admin_teacher_export_endpoint_sanitizes_download_filename(tmp_path: Path):
    db_path, _evaluation_id, teacher_id = closed_evaluation_with_aggregates(tmp_path)
    with connect(db_path) as db:
        db.execute("update evaluations set school_year = ?", ('2025/"26',))
    client = TestClient(create_app(db_path=db_path, admin_password="secret"))

    login(client)
    response = client.get(f"/admin/teachers/{teacher_id}/2025%2F%2226/export.xlsx")

    assert response.status_code == 200
    assert response.headers["content-disposition"] == (
        'attachment; filename="teacher-1-2025-_26.xlsx"'
    )


def closed_evaluation_with_aggregates(tmp_path: Path) -> tuple[Path, int, int]:
    db_path = tmp_path / "eval.db"
    initialize_database(db_path)
    teacher_id = insert_teacher(
        db_path, name="Frau Mueller", email="mueller@example.test"
    )
    evaluation_id = insert_evaluation(
        db_path,
        teacher_id=teacher_id,
        class_group="8b",
        subject="Mathematik",
        student_token="student-token-1",
        teacher_token="teacher-token-1",
    )
    insert_aggregate(db_path, evaluation_id, "Item 1", joined_count=3)
    return db_path, evaluation_id, teacher_id


def insert_teacher(db_path: Path, *, name: str, email: str) -> int:
    with connect(db_path) as db:
        cursor = db.execute(
            "insert into teachers (name, email) values (?, ?)", (name, email)
        )
        return int(cursor.lastrowid)


def insert_evaluation(
    db_path: Path,
    *,
    teacher_id: int,
    class_group: str,
    subject: str,
    student_token: str,
    teacher_token: str,
    status: str = "closed",
) -> int:
    with connect(db_path) as db:
        cursor = db.execute(
            """
            insert into evaluations (
                teacher_id,
                title,
                school_year,
                grade,
                class_group,
                subject,
                questionnaire_version,
                expected_participants,
                status,
                student_token,
                teacher_token,
                updated_at
            ) values (?, ?, '2025/26', 8, ?, ?, 'oberstufe-v1', 24,
                ?, ?, ?, '2026-05-15 14:30:00')
            """,
            (
                teacher_id,
                f"{class_group} {subject}",
                class_group,
                subject,
                status,
                student_token,
                teacher_token,
            ),
        )
        return int(cursor.lastrowid)


def insert_aggregate(
    db_path: Path, evaluation_id: int, item_key: str, *, joined_count: int
) -> None:
    with connect(db_path) as db:
        db.execute(
            """
            insert into item_aggregates (
                evaluation_id,
                item_key,
                count_0,
                count_1,
                count_2,
                count_3,
                missing_count,
                joined_count,
                mean
            ) values (?, ?, 1, 0, 1, 1, 0, ?, 2.0)
            """,
            (evaluation_id, item_key, joined_count),
        )


def login(client: TestClient) -> None:
    response = client.post(
        "/admin/login", data={"password": "secret"}, follow_redirects=False
    )
    assert response.status_code == 303
