from io import BytesIO, StringIO
from pathlib import Path

from fastapi.testclient import TestClient
from openpyxl import load_workbook

from lehr_eval.app import create_app
from lehr_eval.imports import import_master_data
from lehr_eval.migrations import initialize_database


CSV = """schuljahr,klassenstufe,klasse_lerngruppe,fach,lehrkraft_name,lehrkraft_kennung,erwartete_teilnehmerzahl
2025/26,8,8b,Mathematik,Frau Mueller,mueller@example.edu,1
"""


def test_import_activate_join_answer_close_export_flow(tmp_path: Path):
    db_path = tmp_path / "eval.db"
    initialize_database(db_path)
    result = import_master_data(
        db_path, StringIO(CSV), base_url="https://eval.schule.test"
    )
    qr = result.qr_rows[0]

    app = create_app(db_path=db_path, admin_password="secret")
    admin = TestClient(app)
    teacher = TestClient(app)
    student = TestClient(app)

    admin_login = admin.post(
        "/admin/login", data={"password": "secret"}, follow_redirects=False
    )
    activate = admin.post(
        f"/admin/evaluations/{qr.evaluation_id}/activate",
        follow_redirects=False,
    )

    join = student.get(qr.student_path)
    teacher_login = teacher.post(qr.teacher_path, data={"pin": qr.teacher_pin})
    start_joining = teacher.post(
        f"/teacher/{qr.evaluation_id}/start", follow_redirects=False
    )
    show_first_item = teacher.post(
        f"/teacher/{qr.evaluation_id}/item", follow_redirects=False
    )
    open_answers = teacher.post(
        f"/teacher/{qr.evaluation_id}/answers", follow_redirects=False
    )
    answer = student.post(
        f"{qr.student_path}/answer", data={"value": "3"}, follow_redirects=False
    )
    close = teacher.post(
        f"/teacher/{qr.evaluation_id}/close", follow_redirects=False
    )
    export = admin.get(f"/admin/evaluations/{qr.evaluation_id}/export.xlsx")

    assert result.created_evaluations == 1
    assert admin_login.status_code == 303
    assert activate.status_code == 303
    assert join.status_code == 200
    assert "Dein Tiername" in join.text
    assert teacher_login.status_code == 200
    assert "Live-Steuerung" in teacher_login.text
    assert start_joining.status_code == 303
    assert show_first_item.status_code == 303
    assert open_answers.status_code == 303
    assert answer.status_code == 303
    assert close.status_code == 303
    assert export.status_code == 200
    assert export.headers["content-type"] == (
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )

    workbook = load_workbook(BytesIO(export.content))
    sheet = workbook.active
    rows = list(sheet.iter_rows(values_only=True))
    aggregate = next(row for row in rows if row and row[0] == "Item 1")
    assert aggregate[:8] == ("Item 1", 0, 0, 0, 1, 0, 1, 3.0)
