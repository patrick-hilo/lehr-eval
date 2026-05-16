from __future__ import annotations

from dataclasses import dataclass
import csv
import hashlib
import io
import secrets
from pathlib import Path
from typing import BinaryIO, TextIO

from openpyxl import load_workbook

from lehr_eval.db import connect
from lehr_eval.questionnaires import questionnaire_for_grade


REQUIRED_COLUMNS = (
    "schuljahr",
    "klassenstufe",
    "klasse_lerngruppe",
    "fach",
    "lehrkraft_name",
    "lehrkraft_kennung",
    "erwartete_teilnehmerzahl",
)


class ImportErrorReport(Exception):
    def __init__(self, errors: list[str]):
        self.errors = errors
        super().__init__("\n".join(errors))


@dataclass(frozen=True)
class ImportResult:
    created_evaluations: int
    qr_rows: list["QrRow"]


@dataclass(frozen=True)
class QrRow:
    evaluation_id: int
    student_path: str
    teacher_path: str
    student_url: str
    teacher_url: str
    teacher_pin: str | None
    school_year: str
    class_group: str
    subject: str
    teacher_name: str


@dataclass(frozen=True)
class _ImportRow:
    row_number: int
    school_year: str
    grade: int
    class_group: str
    subject: str
    teacher_name: str
    teacher_identifier: str
    expected_participants: int
    questionnaire_version: str


def import_master_data(
    db_path: str | Path, file_obj: TextIO, base_url: str
) -> ImportResult:
    rows = _read_and_validate_csv(file_obj)
    return _persist_rows(db_path, rows, base_url=base_url)


def import_master_data_from_xlsx(
    db_path: str | Path, file_obj: BinaryIO, base_url: str
) -> ImportResult:
    rows = _read_and_validate_xlsx(file_obj)
    return _persist_rows(db_path, rows, base_url=base_url)


def _persist_rows(
    db_path: str | Path, rows: list["_ImportRow"], *, base_url: str
) -> ImportResult:
    _validate_duplicate_groups(db_path, rows)

    qr_rows: list[QrRow] = []
    current_import_pins: dict[tuple[str, str], str | None] = {}
    with connect(db_path) as db:
        for row in rows:
            teacher_id = _upsert_teacher(
                db, name=row.teacher_name, identifier=row.teacher_identifier
            )
            teacher_pin = _pin_for_import_row(
                db,
                teacher_id=teacher_id,
                teacher_identifier=row.teacher_identifier,
                school_year=row.school_year,
                current_import_pins=current_import_pins,
            )

            student_token = secrets.token_urlsafe(16)
            teacher_token = secrets.token_urlsafe(16)
            title = f"{row.class_group} {row.subject}"
            cursor = db.execute(
                """
                insert into evaluations (
                    teacher_id,
                    title,
                    school_year,
                    grade,
                    class_group,
                    subject,
                    questionnaire_version,
                    expected_participants,
                    base_url,
                    status,
                    student_token,
                    teacher_token
                ) values (?, ?, ?, ?, ?, ?, ?, ?, ?, 'prepared', ?, ?)
                """,
                (
                    teacher_id,
                    title,
                    row.school_year,
                    row.grade,
                    row.class_group,
                    row.subject,
                    row.questionnaire_version,
                    row.expected_participants,
                    base_url.rstrip("/"),
                    student_token,
                    teacher_token,
                ),
            )
            evaluation_id = int(cursor.lastrowid)
            student_path = f"/e/{student_token}"
            teacher_path = f"/t/{teacher_token}"
            qr_rows.append(
                QrRow(
                    evaluation_id=evaluation_id,
                    student_path=student_path,
                    teacher_path=teacher_path,
                    student_url=_join_url(base_url, student_path),
                    teacher_url=_join_url(base_url, teacher_path),
                    teacher_pin=teacher_pin,
                    school_year=row.school_year,
                    class_group=row.class_group,
                    subject=row.subject,
                    teacher_name=row.teacher_name,
                )
            )

    return ImportResult(created_evaluations=len(qr_rows), qr_rows=qr_rows)


def _read_and_validate_xlsx(file_obj: BinaryIO) -> list[_ImportRow]:
    try:
        workbook = load_workbook(file_obj, data_only=True, read_only=True)
    except Exception as error:
        raise ImportErrorReport(
            [f"Excel-Datei konnte nicht gelesen werden: {error}"]
        ) from error

    sheet = workbook.active
    if sheet is None:
        raise ImportErrorReport(["Excel-Datei enthaelt keine Tabelle"])

    rows_iter = sheet.iter_rows(values_only=True)
    try:
        header = next(rows_iter)
    except StopIteration:
        raise ImportErrorReport(["Excel-Datei enthaelt keine Kopfzeile"])

    csv_buffer = io.StringIO()
    writer = csv.writer(csv_buffer)
    writer.writerow([_cell_text(cell) for cell in header])
    for raw_row in rows_iter:
        if all(cell is None or _cell_text(cell) == "" for cell in raw_row):
            continue
        writer.writerow([_cell_text(cell) for cell in raw_row])

    csv_buffer.seek(0)
    return _read_and_validate_csv(csv_buffer)


def _cell_text(value) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def _read_and_validate_csv(file_obj: TextIO) -> list[_ImportRow]:
    reader = csv.DictReader(file_obj)
    if reader.fieldnames is None:
        raise ImportErrorReport(["CSV-Datei enthaelt keine Kopfzeile"])

    fieldnames = tuple(name.strip() for name in reader.fieldnames)
    if fieldnames != REQUIRED_COLUMNS:
        missing = [column for column in REQUIRED_COLUMNS if column not in fieldnames]
        extra = [column for column in fieldnames if column not in REQUIRED_COLUMNS]
        details = []
        if missing:
            details.append("fehlende Spalten: " + ", ".join(missing))
        if extra:
            details.append("unerwartete Spalten: " + ", ".join(extra))
        raise ImportErrorReport(["CSV-Spalten ungueltig (" + "; ".join(details) + ")"])

    errors: list[str] = []
    rows: list[_ImportRow] = []
    seen_groups: set[tuple[str, str, str, str]] = set()
    for row_number, raw_row in enumerate(reader, start=2):
        row, row_errors = _validate_row(row_number, raw_row)
        errors.extend(row_errors)
        if row is None:
            continue

        group_key = (
            row.school_year,
            row.class_group,
            row.subject,
            row.teacher_identifier,
        )
        if group_key in seen_groups:
            errors.append(f"Zeile {row_number}: Doppelte Unterrichtsgruppe")
        seen_groups.add(group_key)
        rows.append(row)

    if errors:
        raise ImportErrorReport(errors)
    return rows


def _validate_row(
    row_number: int, raw_row: dict[str, str | None]
) -> tuple[_ImportRow | None, list[str]]:
    errors: list[str] = []
    values = {
        column: (raw_row.get(column) or "").strip() for column in REQUIRED_COLUMNS
    }

    for column, value in values.items():
        if value == "":
            errors.append(f"Zeile {row_number}: {column} ist erforderlich")

    grade = _parse_int(values["klassenstufe"])
    if grade is None or not 1 <= grade <= 10:
        errors.append(f"Zeile {row_number}: Klassenstufe muss zwischen 1 und 10 liegen")

    expected_participants = _parse_int(values["erwartete_teilnehmerzahl"])
    if expected_participants is None or expected_participants < 0:
        errors.append(
            f"Zeile {row_number}: erwartete_teilnehmerzahl muss groesser gleich 0 sein"
        )

    if errors:
        return None, errors

    assert grade is not None
    assert expected_participants is not None
    questionnaire = questionnaire_for_grade(grade)
    return (
        _ImportRow(
            row_number=row_number,
            school_year=values["schuljahr"],
            grade=grade,
            class_group=values["klasse_lerngruppe"],
            subject=values["fach"],
            teacher_name=values["lehrkraft_name"],
            teacher_identifier=values["lehrkraft_kennung"].lower(),
            expected_participants=expected_participants,
            questionnaire_version=questionnaire.version,
        ),
        [],
    )


def _validate_duplicate_groups(db_path: str | Path, rows: list[_ImportRow]) -> None:
    errors: list[str] = []
    with connect(db_path) as db:
        for row in rows:
            existing = db.execute(
                """
                select 1
                from evaluations
                join teachers on teachers.id = evaluations.teacher_id
                where evaluations.school_year = ?
                  and evaluations.class_group = ?
                  and evaluations.subject = ?
                  and lower(teachers.email) = ?
                """,
                (
                    row.school_year,
                    row.class_group,
                    row.subject,
                    row.teacher_identifier,
                ),
            ).fetchone()
            if existing is not None:
                errors.append(f"Zeile {row.row_number}: Doppelte Unterrichtsgruppe")

    if errors:
        raise ImportErrorReport(errors)


def _upsert_teacher(db, *, name: str, identifier: str) -> int:
    db.execute(
        """
        insert into teachers (name, email)
        values (?, ?)
        on conflict(email) do update set name = excluded.name
        """,
        (name, identifier),
    )
    row = db.execute("select id from teachers where email = ?", (identifier,)).fetchone()
    return int(row["id"])


def _pin_for_import_row(
    db,
    *,
    teacher_id: int,
    teacher_identifier: str,
    school_year: str,
    current_import_pins: dict[tuple[str, str], str | None],
) -> str | None:
    pin_key = (teacher_identifier, school_year)
    if pin_key in current_import_pins:
        return current_import_pins[pin_key]

    existing = db.execute(
        """
        select 1 from teacher_pins
        where teacher_id = ? and school_year = ?
        """,
        (teacher_id, school_year),
    ).fetchone()
    if existing is not None:
        current_import_pins[pin_key] = None
        return None

    teacher_pin = _generate_pin()
    db.execute(
        """
        insert into teacher_pins (teacher_id, school_year, pin_code, pin_hash)
        values (?, ?, ?, ?)
        """,
        (teacher_id, school_year, teacher_pin, _hash_pin(teacher_pin)),
    )
    current_import_pins[pin_key] = teacher_pin
    return teacher_pin


def _generate_pin() -> str:
    return f"{secrets.randbelow(10000):04d}"


def _hash_pin(pin: str, *, iterations: int = 200_000) -> str:
    salt = secrets.token_bytes(16)
    pin_hash = hashlib.pbkdf2_hmac(
        "sha256", pin.encode("utf-8"), salt, iterations
    )
    return (
        f"pbkdf2_sha256${iterations}${salt.hex()}${pin_hash.hex()}"
    )


def _parse_int(value: str) -> int | None:
    try:
        return int(value)
    except ValueError:
        return None


def _join_url(base_url: str, path: str) -> str:
    return base_url.rstrip("/") + path
