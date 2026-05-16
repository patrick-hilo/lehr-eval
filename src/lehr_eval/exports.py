from __future__ import annotations

from html import escape
from io import BytesIO
from pathlib import Path
import re
from urllib.parse import quote
from zipfile import ZIP_DEFLATED, ZipFile

from openpyxl import Workbook
from openpyxl.worksheet.worksheet import Worksheet
import qrcode

from lehr_eval.db import connect
from lehr_eval.settings import load_settings


class ExportNotFound(ValueError):
    pass


class ExportNotAvailable(ValueError):
    pass


def build_single_export(db_path: str | Path, evaluation_id: int) -> bytes:
    with connect(db_path) as db:
        evaluation = _evaluation_row(db, evaluation_id)
        aggregates = _aggregate_rows(db, evaluation_id)

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = _excel_sheet_name(
        f"{evaluation['class_group']} {evaluation['subject']}", set()
    )
    _write_evaluation_sheet(sheet, evaluation, aggregates)
    return _workbook_bytes(workbook)


def build_teacher_export(db_path: str | Path, teacher_id: int, school_year: str) -> bytes:
    with connect(db_path) as db:
        evaluations = db.execute(
            """
            select
                evaluations.id,
                evaluations.school_year,
                evaluations.grade,
                evaluations.class_group,
                evaluations.subject,
                evaluations.questionnaire_version,
                evaluations.expected_participants,
                evaluations.status,
                evaluations.updated_at,
                teachers.name as teacher_name
            from evaluations
            join teachers on teachers.id = evaluations.teacher_id
            where evaluations.teacher_id = ?
              and evaluations.school_year = ?
              and evaluations.status = 'closed'
            order by evaluations.id
            """,
            (teacher_id, school_year),
        ).fetchall()
        if not evaluations:
            raise ExportNotFound("no closed evaluations found")
        aggregate_rows = {
            int(evaluation["id"]): _aggregate_rows(db, int(evaluation["id"]))
            for evaluation in evaluations
        }

    workbook = Workbook()
    default_sheet = workbook.active
    workbook.remove(default_sheet)
    used_names: set[str] = set()
    for evaluation in evaluations:
        title = _excel_sheet_name(
            f"{evaluation['class_group']} {evaluation['subject']}", used_names
        )
        sheet = workbook.create_sheet(title=title)
        _write_evaluation_sheet(sheet, evaluation, aggregate_rows[int(evaluation["id"])])

    return _workbook_bytes(workbook)


def build_qr_material_zip(db_path: str | Path, evaluation_ids: list[int]) -> bytes:
    with connect(db_path) as db:
        evaluations = _qr_material_rows(db, evaluation_ids)
    _ensure_printable_teacher_pins(evaluations)

    output = BytesIO()
    with ZipFile(output, "w", compression=ZIP_DEFLATED) as archive:
        html_entries = []
        for evaluation in evaluations:
            evaluation_id = int(evaluation["id"])
            student_filename = f"evaluation-{evaluation_id}-schueler.png"
            teacher_filename = f"evaluation-{evaluation_id}-lehrkraft.png"
            student_url = _evaluation_url(evaluation, "student_token", "e")
            teacher_url = _evaluation_url(evaluation, "teacher_token", "t")

            archive.writestr(student_filename, _qr_png(student_url))
            archive.writestr(teacher_filename, _qr_png(teacher_url))
            html_entries.append(
                _qr_html_entry(
                    evaluation=evaluation,
                    student_filename=student_filename,
                    teacher_filename=teacher_filename,
                    student_url=student_url,
                    teacher_url=teacher_url,
                )
            )

        archive.writestr("qr-material.html", _qr_material_html(html_entries))

    return output.getvalue()


def _evaluation_row(db, evaluation_id: int):
    row = db.execute(
        """
        select
            evaluations.id,
            evaluations.school_year,
            evaluations.grade,
            evaluations.class_group,
            evaluations.subject,
            evaluations.questionnaire_version,
            evaluations.expected_participants,
            evaluations.status,
            evaluations.updated_at,
            teachers.name as teacher_name
        from evaluations
        join teachers on teachers.id = evaluations.teacher_id
        where evaluations.id = ?
        """,
        (evaluation_id,),
    ).fetchone()
    if row is None:
        raise ExportNotFound("evaluation not found")
    if row["status"] != "closed":
        raise ExportNotAvailable("evaluation is not closed")
    return row


def _aggregate_rows(db, evaluation_id: int):
    return db.execute(
        """
        select
            item_key,
            count_0,
            count_1,
            count_2,
            count_3,
            missing_count,
            joined_count,
            mean
        from item_aggregates
        where evaluation_id = ?
        order by id
        """,
        (evaluation_id,),
    ).fetchall()


def _qr_material_rows(db, evaluation_ids: list[int]):
    if not evaluation_ids:
        raise ExportNotFound("no evaluations requested")

    placeholders = ", ".join("?" for _ in evaluation_ids)

    rows = db.execute(
        f"""
        select
            evaluations.id,
            evaluations.school_year,
            evaluations.class_group,
            evaluations.subject,
            evaluations.student_token,
            evaluations.teacher_token,
            evaluations.base_url,
            teachers.name as teacher_name,
            teacher_pins.pin_code as teacher_pin
        from evaluations
        join teachers on teachers.id = evaluations.teacher_id
        left join teacher_pins
            on teacher_pins.teacher_id = evaluations.teacher_id
           and teacher_pins.school_year = evaluations.school_year
        where evaluations.id in ({placeholders})
        order by evaluations.id
        """,
        evaluation_ids,
    ).fetchall()
    if len(rows) != len(set(evaluation_ids)):
        raise ExportNotFound("evaluation not found")
    return rows


def _ensure_printable_teacher_pins(evaluations) -> None:
    missing = [
        str(evaluation["id"])
        for evaluation in evaluations
        if not evaluation["teacher_pin"]
    ]
    if missing:
        raise ExportNotAvailable(
            "teacher pin is not printable for evaluation(s): " + ", ".join(missing)
        )


def _evaluation_url(evaluation, token_column: str, prefix: str) -> str:
    base_url = (evaluation["base_url"] or load_settings().base_url).rstrip("/")
    token = quote(str(evaluation[token_column]), safe="")
    return f"{base_url}/{prefix}/{token}"


def _qr_png(url: str) -> bytes:
    image = qrcode.make(url)
    output = BytesIO()
    image.save(output, format="PNG")
    return output.getvalue()


def _qr_material_html(entries: list[str]) -> str:
    return """<!doctype html>
<html lang="de">
<head>
  <meta charset="utf-8">
  <title>QR-Material</title>
  <style>
    body { font-family: Arial, sans-serif; margin: 24px; color: #111; }
    .evaluation { break-after: page; page-break-after: always; margin-bottom: 32px; }
    .qr-grid { display: grid; grid-template-columns: repeat(2, minmax(0, 1fr)); gap: 24px; }
    img { width: 220px; height: 220px; }
    dt { font-weight: 700; }
    dd { margin: 0 0 8px; }
    .url { word-break: break-all; }
  </style>
</head>
<body>
""" + "\n".join(entries) + """
</body>
</html>
"""


def _qr_html_entry(
    *,
    evaluation,
    student_filename: str,
    teacher_filename: str,
    student_url: str,
    teacher_url: str,
) -> str:
    return f"""  <section class="evaluation">
    <h1>{escape(evaluation["class_group"])} {escape(evaluation["subject"])}</h1>
    <dl>
      <dt>Schuljahr</dt>
      <dd>{escape(evaluation["school_year"])}</dd>
      <dt>Klasse/Lerngruppe</dt>
      <dd>{escape(evaluation["class_group"])}</dd>
      <dt>Fach</dt>
      <dd>{escape(evaluation["subject"])}</dd>
      <dt>Lehrkraft</dt>
      <dd>{escape(evaluation["teacher_name"])}</dd>
      <dt>Lehrkraft-PIN</dt>
      <dd>{escape(evaluation["teacher_pin"] or "")}</dd>
    </dl>
    <div class="qr-grid">
      <section>
        <h2>Schuelerinnen und Schueler</h2>
        <img src="{escape(student_filename)}" alt="QR-Code fuer Schuelerinnen und Schueler">
        <p class="url">{escape(student_url)}</p>
      </section>
      <section>
        <h2>Lehrkraft</h2>
        <img src="{escape(teacher_filename)}" alt="QR-Code fuer die Lehrkraft">
        <p class="url">{escape(teacher_url)}</p>
      </section>
    </div>
  </section>"""


def _write_evaluation_sheet(sheet: Worksheet, evaluation, aggregates) -> None:
    joined_count = _joined_count(aggregates)
    closed_at = evaluation["updated_at"] if evaluation["status"] == "closed" else ""
    rows = (
        ("Schuljahr", _safe_cell_text(evaluation["school_year"])),
        ("Klasse/Lerngruppe", _safe_cell_text(evaluation["class_group"])),
        ("Klassenstufe", int(evaluation["grade"])),
        ("Fach", _safe_cell_text(evaluation["subject"])),
        ("Lehrkraft", _safe_cell_text(evaluation["teacher_name"])),
        ("Fragebogen-Version", _safe_cell_text(evaluation["questionnaire_version"])),
        ("Erwartete Teilnehmerzahl", int(evaluation["expected_participants"])),
        ("tatsaechliche Beitritte", joined_count),
        ("Hinweis: Auswertung enthaelt nur aggregierte Daten", None),
        ("Abschlussdatum", _safe_cell_text(closed_at)),
    )
    for row_index, row in enumerate(rows, start=1):
        sheet.cell(row=row_index, column=1, value=row[0])
        sheet.cell(row=row_index, column=2, value=row[1])

    headers = ("Item", "0", "1", "2", "3", "Fehlend", "Beitritte", "Mittelwert")
    for column_index, header in enumerate(headers, start=1):
        sheet.cell(row=12, column=column_index, value=header)

    for row_index, aggregate in enumerate(aggregates, start=13):
        values = (
            _safe_cell_text(aggregate["item_key"]),
            int(aggregate["count_0"]),
            int(aggregate["count_1"]),
            int(aggregate["count_2"]),
            int(aggregate["count_3"]),
            int(aggregate["missing_count"]),
            int(aggregate["joined_count"]),
            None if aggregate["mean"] is None else float(aggregate["mean"]),
        )
        for column_index, value in enumerate(values, start=1):
            sheet.cell(row=row_index, column=column_index, value=value)


def _joined_count(aggregates) -> int:
    if not aggregates:
        return 0
    return max(int(row["joined_count"]) for row in aggregates)


def _excel_sheet_name(raw_name: str, used_names: set[str]) -> str:
    name = re.sub(r"[\[\]:*?/\\]+", "", raw_name).strip() or "Evaluation"
    base = name[:31]
    candidate = base
    counter = 2
    while candidate.casefold() in used_names:
        suffix = f" {counter}"
        candidate = base[: 31 - len(suffix)] + suffix
        counter += 1
    used_names.add(candidate.casefold())
    return candidate


def _safe_cell_text(value) -> str:
    text = "" if value is None else str(value)
    if text.startswith(("=", "+", "-", "@", "\t", "\r")):
        return f"'{text}"
    return text


def safe_download_filename(value: str) -> str:
    filename = value.replace("/", "-")
    return re.sub(r"[^A-Za-z0-9._-]+", "_", filename).strip("._") or "export"


def _workbook_bytes(workbook: Workbook) -> bytes:
    output = BytesIO()
    workbook.save(output)
    return output.getvalue()
